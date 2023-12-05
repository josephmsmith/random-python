import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

def process_excel(file_path):
    # Load the Excel file into a pandas DataFrame
    df = pd.read_excel(file_path)

    # # Delete column 'E'
    df.drop(df.columns[4], axis=1, inplace=True)  # Assuming E is the 5th column
    
    # # Sort by MATCH column if needed
    df.sort_values(by='MATCH', inplace=True)

    # Save the modified DataFrame back to Excel, without the default index
    df.to_excel(file_path, index=False)

    # # Load the workbook and select the active worksheet
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    # # Define the range for your table (modify as per your data)
    # table_range = "A1:D10"  # Example range, adjust to your data
    # table = openpyxl.worksheet.table.Table(displayName="Table1", ref=table_range)
    # table.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(
    #     name="TableStyleMedium9",  # 'Medium Style 2 Blue'
    #     showFirstColumn=False,
    #     showLastColumn=False,
    #     showRowStripes=False,
    #     showColumnStripes=True
    # )
    # worksheet.add_table(table)

    # # Freeze the first column
    # worksheet.freeze_panes = 'B2'

    # # Add conditional formatting
    # match_column_index = df.columns.get_loc("MATCH") + 1
    # yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    # formula = f'"{get_column_letter(match_column_index)}2=FALSE"'
    # worksheet.conditional_formatting.add(f'{get_column_letter(match_column_index)}2:{get_column_letter(match_column_index)}1048576',
    #                                      FormulaRule(formula=formula, fill=yellow_fill))

    # # Save the changes to the file
    # workbook.save(file_path)

if __name__ == "__main__":
    file_path = input("Enter the path to the Excel file: ")
    process_excel(file_path)
